import docx
from docx import Document
from lxml import etree
from openpyxl import workbook
import pandas as pd
import os
import time, datetime
import xlrd  #读取Excel文件的包
import xlsxwriter   #将文件写入Excel的包
dir=r'C:\Users\pangyuelong\Desktop\备份'
content=[]
def data_content(path):
    doc = Document(path)
    table = doc.tables[3]
    data_1=[]
    for i in range(2,len(table.rows)):
        item={}
        item['杆塔区间']=table.cell(i,1).text
        # item['相位'] = table.cell(i, 2).text
        item['距小号塔距离（米）'] = float(table.cell(i, 2).text)
        item['经纬度'] = table.cell(i, 3).text
        item['缺陷类型'] = table.cell(i, 4).text
        item['缺陷级别'] = table.cell(i, 5).text
        item['水平'] = float(table.cell(i, 6).text)
        item['垂直'] = float(table.cell(i, 7).text)
        item['净空'] = float(table.cell(i, 8).text)
        item['侧视图']=path.split('\\')[-1].split('.docx')[0]+'_image'+str(2*i-2)+'.png'
        item['俯视图']=path.split('\\')[-1].split('.docx')[0]+'_image'+str(2*i-1)+'.png'
        item['线路名称']=path.split('\\')[-1].split('_')[1]
        # data_time=path.split("-")[-1].split('.docx')[0]
        # timeArray = time.strptime(data_time, "%m%d")
        # item["拍摄时间"]="2021年"+str(timeArray.tm_mon)+"月"+str(timeArray.tm_mday)+"日"
        data_1.append(item)
    return data_1
for files in os.listdir(r'C:\Users\pangyuelong\Desktop\备份'):
    if files.split('.')[-1]=='docx':
        file_name=os.path.join(dir,files)
        content.extend(data_content(file_name))
pf = pd.DataFrame(content)
file_path = pd.ExcelWriter('6月15日.xlsx')  # 打开excel文件
# pf.fillna(' ', inplace=True)
pf.to_excel(file_path, encoding='utf-8',sheet_name="sheet1")
file_path.save()
